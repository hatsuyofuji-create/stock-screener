# -*- coding: utf-8 -*-
"""
=====================================================================
 受注一覧エクセルの書き出し
=====================================================================
 各Q単独の受注高（百万円）を、見やすい表にして .xlsx に保存します。
   - シート「受注高_各Q単独」: 銘柄×決算期×セグメント で Q1〜Q4 の単独値
   - 合計行（セグメント名 = 合計）も入れます
=====================================================================
"""

from __future__ import annotations
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from src.transform.quarterly import QuarterlyOrders


_HEADERS = ["銘柄コード", "会社名", "決算期", "セグメント",
            "Q1単独", "Q2単独", "Q3単独", "Q4単独", "単位"]


def _fmt(v):
    return "" if v is None else round(v, 1)


def write_xlsx(results: Iterable[QuarterlyOrders], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "受注高_各Q単独"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    for col, name in enumerate(_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center")

    row = 2
    for r in results:
        # セグメント別
        for seg, single in sorted(r.single_by_segment.items()):
            ws.cell(row=row, column=1, value=r.code)
            ws.cell(row=row, column=2, value=r.company)
            ws.cell(row=row, column=3, value=r.fiscal_year)
            ws.cell(row=row, column=4, value=seg)
            ws.cell(row=row, column=5, value=_fmt(single.get(1)))
            ws.cell(row=row, column=6, value=_fmt(single.get(2)))
            ws.cell(row=row, column=7, value=_fmt(single.get(3)))
            ws.cell(row=row, column=8, value=_fmt(single.get(4)))
            ws.cell(row=row, column=9, value="百万円")
            row += 1
        # 合計
        ws.cell(row=row, column=1, value=r.code)
        ws.cell(row=row, column=2, value=r.company)
        ws.cell(row=row, column=3, value=r.fiscal_year)
        c = ws.cell(row=row, column=4, value="合計")
        c.font = Font(bold=True)
        ws.cell(row=row, column=5, value=_fmt(r.single_total.get(1)))
        ws.cell(row=row, column=6, value=_fmt(r.single_total.get(2)))
        ws.cell(row=row, column=7, value=_fmt(r.single_total.get(3)))
        ws.cell(row=row, column=8, value=_fmt(r.single_total.get(4)))
        ws.cell(row=row, column=9, value="百万円")
        row += 1

    # 列幅をざっくり調整
    widths = [10, 22, 10, 18, 12, 12, 12, 12, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    wb.save(out_path)
