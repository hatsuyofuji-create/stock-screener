"""API 統合テスト（Phase 0/1）。一時 SQLite を使う。"""

import io
import os
import tempfile

import pytest
from openpyxl import Workbook

# app を import する前にテスト用 DB を指定する。
_tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
_tmp.close()
os.environ["DATABASE_URL"] = "sqlite:///" + _tmp.name

from fastapi.testclient import TestClient  # noqa: E402
from app.main import app  # noqa: E402
from app.db import init_db  # noqa: E402


@pytest.fixture(scope="module")
def client():
    init_db()
    with TestClient(app) as c:
        yield c
    os.unlink(_tmp.name)


def test_disclaimer_present(client):
    r = client.get("/")
    assert r.status_code == 200
    assert "投資助言" in r.json()["disclaimer"]


def test_full_flow(client):
    # 企業作成
    r = client.post("/api/companies", json={
        "code": "9999", "name": "テスト商事", "sector": "卸売",
        "growth_type": "安定成長", "moat_tags": ["ブランド", "コスト競争力"],
        "qualitative_note": "テスト用の定性メモ",
    })
    assert r.status_code == 200
    assert r.json()["is_cyclical"] is False

    # 前期・最新期を投入
    prior = {
        "fiscal_period": "2024-03", "revenue": 1000, "cogs": 600,
        "operating_income": 100, "ordinary_income": 90, "net_income": 60,
        "operating_cf": 80, "total_assets": 1000, "equity": 500,
        "current_assets": 400, "current_liabilities": 200,
        "interest_bearing_debt": 200, "cash": 100, "capital_stock": 100,
    }
    latest = {
        "fiscal_period": "2025-03", "revenue": 1200, "cogs": 680,
        "operating_income": 150, "ordinary_income": 140, "net_income": 95,
        "operating_cf": 160, "total_assets": 1100, "equity": 600,
        "current_assets": 500, "current_liabilities": 220,
        "interest_bearing_debt": 180, "cash": 150, "capital_stock": 100,
    }
    assert client.post("/api/companies/9999/statements", json=prior).status_code == 200
    assert client.post("/api/companies/9999/statements", json=latest).status_code == 200

    # 指標
    r = client.get("/api/companies/9999/metrics")
    assert r.status_code == 200
    body = r.json()
    assert len(body["metrics"]) == 2
    assert body["changes"] is not None

    # Fスコア
    r = client.get("/api/companies/9999/fscore")
    assert r.status_code == 200
    fs = r.json()
    assert fs["total"] == 9
    assert fs["is_healthy"] is True


def test_fscore_requires_two_periods(client):
    client.post("/api/companies", json={"code": "8888", "name": "一期のみ"})
    client.post("/api/companies/8888/statements", json={
        "fiscal_period": "2025-03", "net_income": 10, "operating_cf": 20,
    })
    r = client.get("/api/companies/8888/fscore")
    assert r.status_code == 400


def test_cyclical_flag(client):
    client.post("/api/companies", json={
        "code": "7777", "name": "シクリカル鋼業", "growth_type": "シクリカル",
    })
    r = client.get("/api/companies/7777")
    assert r.json()["is_cyclical"] is True


def test_valuation_endpoint(client):
    """Phase 2: 適正株価・安全域・リスクリワードの一括計算。"""
    payload = {
        "current_price": 10,
        "multiples": {"fair_per": 15, "fair_pbr": 1.5, "fair_yield": 0.03},
        "base": {
            "scenario": "base", "revenue": 1000, "operating_margin": 0.15,
            "payout_ratio": 0.30, "shares_outstanding": 100, "prior_bps": 5.0,
        },
        "negative": {
            "scenario": "negative", "revenue": 1000, "operating_margin": 0.15,
            "payout_ratio": 0.30, "one_time_items": 50, "shares_outstanding": 100,
            "prior_bps": 5.0,
        },
    }
    r = client.post("/api/valuation", json=payload)
    assert r.status_code == 200
    body = r.json()
    assert abs(body["base"]["projection"]["eps"] - 0.975) < 1e-9
    assert abs(body["base"]["fair_value"]["per"]["fair_price"] - 14.625) < 1e-9
    assert body["risk_reward"]["verdict"] == "損小利大"


def test_company_valuation_saves_forecast(client):
    client.post("/api/companies", json={"code": "6666", "name": "予想保存テスト"})
    payload = {
        "current_price": 10, "save": True, "fiscal_period": "2026-03",
        "multiples": {"fair_per": 15},
        "base": {"revenue": 1000, "operating_margin": 0.15, "payout_ratio": 0.3,
                 "shares_outstanding": 100, "prior_bps": 5.0},
        "negative": {"scenario": "negative", "revenue": 1000, "operating_margin": 0.15,
                     "payout_ratio": 0.3, "one_time_items": 50,
                     "shares_outstanding": 100, "prior_bps": 5.0},
    }
    r = client.post("/api/companies/6666/valuation", json=payload)
    assert r.status_code == 200
    assert r.json().get("saved") is True


def _excel_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "予想"
    ws["A1"] = "前提"
    ws["B1"] = "国内堅調"
    ws["A2"] = "売上高"
    ws["B2"] = 1200
    ws["A3"] = "営業利益"
    ws["B3"] = 180
    ws["A4"] = "予想EPS"
    ws["B4"] = 1.2
    ws["A5"] = "予想BPS"
    ws["B5"] = 20.0
    ws["A6"] = "予想DPS"
    ws["B6"] = 0.36
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_excel_preview_and_extract(client):
    """Phase 3: Excel プレビュー → 自動検出 → 抽出 → Forecast 保存。"""
    client.post("/api/companies", json={"code": "5555", "name": "Excel取込テスト"})

    files = {"file": ("f.xlsx", _excel_bytes(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = client.post("/api/excel/preview", files=files)
    assert r.status_code == 200
    body = r.json()
    file_id = body["file_id"]
    sug = body["suggestions"]
    assert sug["eps"]["value"] == 1.2

    mapping = {f: {"sheet": sug[f]["sheet"], "cell": sug[f]["value_cell"]}
               for f in ("eps", "bps", "dps", "revenue", "operating_income", "qualitative_note")}
    r = client.post("/api/companies/5555/excel/extract", json={
        "file_id": file_id, "mapping": mapping, "fiscal_period": "2026-03",
        "save": True, "save_profile_as": "park24_annual",
    })
    assert r.status_code == 200
    vals = r.json()["values"]
    assert vals["eps"] == 1.2 and vals["bps"] == 20.0 and vals["dps"] == 0.36
    assert r.json()["saved"] is True

    # 定性メモが Excel の「前提」で上書きされている
    assert "国内堅調" in client.get("/api/companies/5555").json()["qualitative_note"]

    # マッピングが保存され再利用できる
    profiles = client.get("/api/mapping_profiles").json()
    assert any(p["format_name"] == "park24_annual" for p in profiles)


def test_forecast_fscore_endpoint(client):
    """Phase 3→4.2 接続：最新実績に予想値を上書きした予想Fスコア。"""
    client.post("/api/companies", json={"code": "4444", "name": "予想Fスコア"})
    prior = {"fiscal_period": "2024-03", "revenue": 1000, "operating_income": 100,
             "ordinary_income": 90, "net_income": 60, "operating_cf": 80,
             "total_assets": 1000, "equity": 500, "current_assets": 400,
             "current_liabilities": 200, "interest_bearing_debt": 200, "capital_stock": 100}
    latest = {"fiscal_period": "2025-03", "revenue": 1100, "operating_income": 110,
              "ordinary_income": 100, "net_income": 70, "operating_cf": 90,
              "total_assets": 1050, "equity": 560, "current_assets": 450,
              "current_liabilities": 210, "interest_bearing_debt": 190, "capital_stock": 100}
    client.post("/api/companies/4444/statements", json=prior)
    client.post("/api/companies/4444/statements", json=latest)

    # 予想で純利益を赤字化 → 項目1 が No になる
    r = client.post("/api/companies/4444/fscore/forecast",
                    json={"net_income": -10, "operating_income": 20})
    assert r.status_code == 200
    body = r.json()
    assert body["mode"] == "forecast"
    item1 = next(i for i in body["items"] if i["number"] == 1)
    assert item1["passed"] is False


def test_screening_and_matrix_flow(client):
    """Phase 4: 市場データ登録 → 閾値/魔法の公式スクリーニング → マトリクス。"""
    # 本命銘柄: 低PBR・高Fスコア・ネットキャッシュ
    client.post("/api/companies", json={"code": "3333", "name": "本命工業"})
    prior = {"fiscal_period": "2024-03", "revenue": 1000, "operating_income": 100,
             "ordinary_income": 90, "net_income": 60, "operating_cf": 80, "investing_cf": -20,
             "total_assets": 1000, "equity": 800, "current_assets": 400,
             "current_liabilities": 100, "interest_bearing_debt": 50, "cash": 300,
             "capital_stock": 100, "shares_outstanding": 100, "dps": 30}
    latest = {"fiscal_period": "2025-03", "revenue": 1200, "operating_income": 150,
              "ordinary_income": 140, "net_income": 95, "operating_cf": 160, "investing_cf": -30,
              "total_assets": 1100, "equity": 900, "current_assets": 500,
              "current_liabilities": 110, "interest_bearing_debt": 40, "cash": 350,
              "capital_stock": 100, "shares_outstanding": 100, "dps": 40}
    client.post("/api/companies/3333/statements", json=prior)
    client.post("/api/companies/3333/statements", json=latest)
    # BPS = 900/100 = 9.0、株価5 → PBR=0.556（低PBR）、利回り=40/5=8
    client.put("/api/companies/3333/market", json={"price": 5.0, "price_3y_ago": 8.0})

    # 閾値: ネットキャッシュ＋FCF>0＋Fスコア7以上
    r = client.post("/api/screen/threshold", json={
        "net_cash_required": True, "fcf_positive": True, "fscore_min": 7,
    })
    assert r.status_code == 200
    assert any(h["code"] == "3333" for h in r.json()["hits"])

    # 魔法の公式
    r = client.post("/api/screen/magic?top_n=10")
    assert r.status_code == 200
    assert any(row["code"] == "3333" for row in r.json()["results"])

    # マトリクス: 本命ゾーン判定
    r = client.get("/api/matrix")
    assert r.status_code == 200
    pt = next(p for p in r.json()["points"] if p["code"] == "3333")
    assert pt["quadrant"] == "honmei"

    # スクリーニング設定の保存
    r = client.post("/api/screen/settings", json={
        "name": "ネットキャッシュ小型", "mode": "threshold",
        "params": {"net_cash_required": True, "fscore_min": 7},
    })
    assert r.json()["saved"] is True
    assert any(s["name"] == "ネットキャッシュ小型" for s in client.get("/api/screen/settings").json())


def test_business_cycle_endpoint(client):
    r = client.get("/api/business-cycle/phases")
    assert len(r.json()["phases"]) == 8
    r = client.post("/api/business-cycle", json={"phase": "順調な拡大"})
    assert r.json()["recommended_method"] == "per"
    r = client.post("/api/business-cycle", json={"phase": "順調な拡大", "is_cyclical": True})
    assert r.json()["recommended_method"] == "pbr"


def test_portfolio_holdings_and_alerts(client):
    client.post("/api/companies", json={
        "code": "2222", "name": "保有テスト", "sector": "機械", "growth_type": "安定成長",
    })
    client.put("/api/companies/2222/market", json={"price": 130})
    r = client.post("/api/holdings", json={
        "company_code": "2222", "buy_price": 100, "shares": 100,
        "buy_date": "2022-01-01", "confidence": 4,
    })
    hid = r.json()["id"]

    # 利確: 現在株価130 ≥ 適正株価100×1.3 → ポジション調整
    r = client.post(f"/api/holdings/{hid}/take-profit", json={"fair_price": 100})
    assert r.json()["action"] == "ポジション調整"

    # 損切り: buy_date 2022 で含み損なら2年超だが、現在株価130>buy100 で含み損でない → 発火せず
    r = client.post(f"/api/holdings/{hid}/stop-loss", json={"current_price": 80})
    assert r.json()["triggered"] is True  # 80<100 かつ2年超

    # 分散チェック
    r = client.get("/api/portfolio/diversification")
    assert r.status_code == 200
    assert r.json()["count"] >= 1

    # ポジションサイズ・再評価
    assert client.post("/api/portfolio/position-size",
                       json={"confidence": 5, "upside": 0.3}).status_code == 200
    r = client.post("/api/portfolio/reevaluate", json={"margin_of_safety": 0.2, "fscore": 8})
    assert r.json()["would_buy"] is True
