"""予想Excel 取り込みのテスト（仕様書 4.6 / 受け入れ基準）。

様式の異なる2ファイル（年次型・四半期型）から予想EPS/BPS/DPS を正しく抽出できること。
"""

import io

import pytest
from openpyxl import Workbook

from app.engines.excel_adapter import (
    load_sheets, auto_detect_mappings, extract_values,
)


def _annual_book() -> bytes:
    """Park24 型（年次トップダウン）：見出しの右隣に値。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "年次予想"
    ws["A1"] = "前提"
    ws["B1"] = "国内回復、海外は横ばい"
    ws["A3"] = "売上高"
    ws["B3"] = 3400
    ws["A4"] = "営業利益"
    ws["B4"] = 408
    ws["A5"] = "営業利益率"
    ws["B5"] = 0.12
    ws["A6"] = "当期純利益"
    ws["B6"] = 265
    ws["A7"] = "予想EPS"
    ws["B7"] = 2.65
    ws["A8"] = "予想BPS"
    ws["B8"] = 30.0
    ws["A9"] = "予想DPS"
    ws["B9"] = 0.80
    ws["A10"] = "ROE"
    ws["B10"] = 0.088
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _quarterly_book() -> bytes:
    """セーレン型（四半期ボトムアップ）：見出しの下に値。別シート構成。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "四半期"
    # 見出し行 → 下に通期予想値
    ws["B1"] = "売上高"
    ws["B2"] = 1200
    ws["C1"] = "営業利益"
    ws["C2"] = 96
    ws["D1"] = "EPS"
    ws["D2"] = 1.50
    ws["E1"] = "BPS"
    ws["E2"] = 18.0
    ws["F1"] = "1株当たり配当"
    ws["F2"] = 0.45
    ws2 = wb.create_sheet("メモ")
    ws2["A1"] = "前提"
    ws2["A2"] = "受注残が下支え"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_annual_auto_detect_and_extract():
    sheets = load_sheets(_annual_book())
    sug = auto_detect_mappings(sheets)
    # EPS/BPS/DPS の値セルが検出できていること
    assert sug["eps"]["value"] == 2.65
    assert sug["bps"]["value"] == 30.0
    assert sug["dps"]["value"] == 0.80
    assert sug["revenue"]["value"] == 3400

    mapping = {f: {"sheet": sug[f]["sheet"], "cell": sug[f]["value_cell"]} for f in sug}
    values = extract_values(sheets, mapping)
    assert values["eps"] == 2.65
    assert values["bps"] == 30.0
    assert values["dps"] == 0.80
    assert values["revenue"] == 3400.0
    # 「前提」欄は定性メモとして取り込む
    assert "国内回復" in values["qualitative_note"]


def test_quarterly_auto_detect_and_extract():
    sheets = load_sheets(_quarterly_book())
    sug = auto_detect_mappings(sheets)
    # 見出しの下に値がある様式でも検出できる
    assert sug["eps"]["value"] == 1.50
    assert sug["bps"]["value"] == 18.0
    assert sug["dps"]["value"] == 0.45

    mapping = {f: {"sheet": sug[f]["sheet"], "cell": sug[f]["value_cell"]} for f in sug}
    values = extract_values(sheets, mapping)
    assert values["eps"] == 1.50
    assert values["bps"] == 18.0
    assert values["dps"] == 0.45
    # 別シートの「前提」も拾う
    assert "受注残" in values["qualitative_note"]


def test_extract_respects_explicit_mapping():
    """自動検出を無視し、ユーザー指定のセルを優先すること。"""
    sheets = load_sheets(_annual_book())
    mapping = {"eps": {"sheet": "年次予想", "cell": "B8"}}  # あえて BPS のセルを指定
    values = extract_values(sheets, mapping)
    assert values["eps"] == 30.0


def test_extract_handles_string_number():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "予想EPS"
    ws["B1"] = "1,234"
    buf = io.BytesIO()
    wb.save(buf)
    sheets = load_sheets(buf.getvalue())
    values = extract_values(sheets, {"eps": {"sheet": ws.title, "cell": "B1"}})
    assert values["eps"] == 1234.0
