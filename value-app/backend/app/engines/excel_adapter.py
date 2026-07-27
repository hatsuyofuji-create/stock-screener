"""予想Excel 取り込みアダプター（仕様書 4.6）。

ユーザーの予想Excel（Park24型の年次トップダウン、セーレン型の四半期ボトムアップ等）は
様式がバラバラなので固定テンプレートを強制しない。マッピング層を実装する。

- load_preview: シートをプレビュー用の2次元グリッドに読み出す。
- auto_detect_mappings: 見出し（EPS 等）を自動検出し、値セルの候補を提案する。
- extract_values: ユーザー確定のマッピングに従い、Forecast に必要な少数の値を抽出する。
- Excel 内の「前提」欄などの自由記述テキストは qualitative_note として取り込む。

抽出対象（Forecast の入力）：予想EPS・予想BPS・予想DPS ＋ 売上・営業利益・ROE、
および定性メモ（qualitative_note）。
"""

from __future__ import annotations

import io
from dataclasses import dataclass, asdict
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, coordinate_to_tuple

# 抽出対象の正規フィールドと、見出し自動検出用キーワード（正規化後に部分一致）。
FIELD_KEYWORDS: dict[str, list[str]] = {
    "eps": ["予想EPS", "EPS", "1株当たり当期純利益", "一株当たり当期純利益", "1株当たり利益"],
    "bps": ["予想BPS", "BPS", "1株当たり純資産", "一株当たり純資産"],
    "dps": ["予想DPS", "DPS", "1株当たり配当", "一株当たり配当", "配当金", "予想配当"],
    "revenue": ["売上高", "営業収益", "売上収益", "売上"],
    "operating_income": ["営業利益"],
    "operating_margin": ["営業利益率"],
    "net_income": ["親会社株主に帰属する当期純利益", "当期純利益", "純利益"],
    "roe": ["ROE", "自己資本利益率"],
    "qualitative_note": ["前提条件", "前提", "シナリオ前提", "コメント", "メモ"],
}

# 数値として扱うフィールド（qualitative_note 以外）
_NUMERIC_FIELDS = {f for f in FIELD_KEYWORDS if f != "qualitative_note"}


def _normalize(text: str) -> str:
    """全角スペース・空白除去 ＋ ラテン文字を大文字化して比較しやすくする。"""
    return "".join(text.split()).replace("　", "").upper()


@dataclass
class SheetData:
    name: str
    cells: dict[str, object]  # "B5" -> 値
    max_row: int
    max_col: int


def load_sheets(data: bytes) -> list[SheetData]:
    """xlsx バイト列から全シートを読み込む（数式は計算済み値を使用）。"""
    wb = load_workbook(io.BytesIO(data), data_only=True)
    sheets: list[SheetData] = []
    for ws in wb.worksheets:
        cells: dict[str, object] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = cell.value
        sheets.append(SheetData(
            name=ws.title, cells=cells,
            max_row=ws.max_row or 0, max_col=ws.max_column or 0,
        ))
    return sheets


def load_preview(data: bytes, max_rows: int = 40, max_cols: int = 15) -> dict:
    """プレビュー用に各シートを2次元グリッド（セル座標＋値）で返す。"""
    sheets = load_sheets(data)
    preview: dict[str, list] = {}
    for sd in sheets:
        rows = min(sd.max_row, max_rows)
        cols = min(sd.max_col, max_cols)
        grid = []
        for r in range(1, rows + 1):
            row_cells = []
            for c in range(1, cols + 1):
                coord = f"{get_column_letter(c)}{r}"
                row_cells.append({"cell": coord, "value": sd.cells.get(coord)})
            grid.append(row_cells)
        preview[sd.name] = grid
    return preview


def _neighbor_value_cell(sd: SheetData, coord: str, want_numeric: bool) -> Optional[str]:
    """見出しセルの隣（右→下）で、望む型の値が入っているセル座標を返す。"""
    row, col = coordinate_to_tuple(coord)
    candidates = [
        f"{get_column_letter(col + 1)}{row}",  # 右
        f"{get_column_letter(col)}{row + 1}",  # 下
        f"{get_column_letter(col + 2)}{row}",  # 右2つ目（単位列を挟む様式）
    ]
    for cand in candidates:
        val = sd.cells.get(cand)
        if val is None:
            continue
        is_num = isinstance(val, (int, float)) and not isinstance(val, bool)
        if want_numeric and is_num:
            return cand
        if not want_numeric and not is_num:
            return cand
    return None


@dataclass
class MappingSuggestion:
    field: str
    sheet: str
    label_cell: str          # 見出しが見つかったセル
    value_cell: Optional[str]  # 値が入っていると推定したセル
    value: object            # 推定値（確認用）

    def to_dict(self) -> dict:
        return asdict(self)


def auto_detect_mappings(sheets: list[SheetData]) -> dict[str, dict]:
    """見出しを走査し、各フィールドの値セル候補を提案する（仕様書 4.6）。

    同一フィールドで複数ヒットした場合は、値セルを特定できた最初の候補を採用する。
    """
    suggestions: dict[str, MappingSuggestion] = {}
    for sd in sheets:
        for coord, val in sd.cells.items():
            if not isinstance(val, str):
                continue
            norm = _normalize(val)
            if not norm:
                continue
            for field, keywords in FIELD_KEYWORDS.items():
                if field in suggestions and suggestions[field].value_cell is not None:
                    continue  # 確定済みの候補があればスキップ
                for kw in keywords:
                    if _normalize(kw) in norm:
                        want_numeric = field in _NUMERIC_FIELDS
                        vcell = _neighbor_value_cell(sd, coord, want_numeric)
                        cur = suggestions.get(field)
                        # 値セルを特定できた候補を優先
                        if cur is None or (cur.value_cell is None and vcell is not None):
                            suggestions[field] = MappingSuggestion(
                                field=field, sheet=sd.name, label_cell=coord,
                                value_cell=vcell,
                                value=sd.cells.get(vcell) if vcell else None,
                            )
                        break
    return {f: s.to_dict() for f, s in suggestions.items()}


def extract_values(sheets: list[SheetData], mapping: dict[str, dict]) -> dict[str, object]:
    """ユーザー確定のマッピングに従い値を抽出する（仕様書 4.6）。

    mapping 例: {"eps": {"sheet": "Sheet1", "cell": "C5"}, ...}
    数値フィールドは float 変換を試み、失敗時は None。qualitative_note は文字列のまま。
    """
    by_name = {sd.name: sd for sd in sheets}
    out: dict[str, object] = {}
    for field, ref in mapping.items():
        sheet_name = ref.get("sheet")
        cell = ref.get("cell")
        sd = by_name.get(sheet_name) if sheet_name else (sheets[0] if sheets else None)
        if sd is None or not cell:
            out[field] = None
            continue
        raw = sd.cells.get(cell)
        if field in _NUMERIC_FIELDS:
            out[field] = _to_float(raw)
        else:
            out[field] = str(raw) if raw is not None else None
    return out


def _to_float(raw) -> Optional[float]:
    if raw is None:
        return None
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return float(raw)
    try:
        # "1,234" や "12.3%" のような文字列に緩く対応
        s = str(raw).replace(",", "").replace("%", "").strip()
        return float(s)
    except (ValueError, TypeError):
        return None
